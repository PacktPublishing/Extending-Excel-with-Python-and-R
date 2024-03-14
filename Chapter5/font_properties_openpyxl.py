# OpenPyXL example for setting font properties
from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
ws = wb.active

# Applying font properties
font = Font(size=14, bold=True, italic=True, color='0000FF')
ws['A1'].font = font

ws['A1'] = 'Name'
ws['B1'] = 'Age'
ws['C1'] = 'City'

wb.save('styled_table_openpyxl.xlsx')
