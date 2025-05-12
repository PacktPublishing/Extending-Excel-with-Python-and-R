# OpenPyXL example for aligning text within cells
from openpyxl import Workbook
from openpyxl.styles import Alignment

wb = Workbook()
ws = wb.active

# Applying text alignment
alignment = Alignment(horizontal='center', vertical='center')
ws['A1'].alignment = alignment

ws['A1'] = 'Name'
ws['B1'] = 'Age'
ws['C1'] = 'City'

wb.save('aligned_table_openpyxl.xlsx')
