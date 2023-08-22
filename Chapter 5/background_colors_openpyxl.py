# openpyxl example for cell background colors
from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb = Workbook()
ws = wb.active

# Applying cell background colors
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
ws['A1'].fill = yellow_fill

ws['A1'] = 'Name'
ws['B1'] = 'Age'
ws['C1'] = 'City'

wb.save('colored_table_openpyxl.xlsx')
