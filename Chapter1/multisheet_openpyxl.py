from openpyxl import load_workbook
import pandas as pd

def read_single_sheet(workbook, sheet_name):
        
        # Load the sheet from the workbook
        sheet = workbook[sheet_name]

        # Read out the raaw data including headers
        sheet_data_raw = sheet.values

        # Separate the headers into a variable
        columns = next(sheet_data_raw)[0:]

        # Create a DataFrame based on the second and subsequent lines of data with the header as column names and return it
        return pd.DataFrame(sheet_data_raw, columns=columns)


def read_multiple_sheets(file_path):

    # Load the workbook
    workbook = load_workbook(file_path)

    # Get a list of all sheet names in the workbook
    sheet_names = workbook.sheetnames

    # Cycle through the sheet names, load the data for each and concatenate them into a single DataFrame
    return pd.concat([read_single_sheet(workbook=workbook, sheet_name=sheet_name) for sheet_name in sheet_names], ignore_index=True)

# Define the file path and sheet names
file_path = 'iris_data.xlsx'

# Read the data from multiple sheets
consolidated_data = read_multiple_sheets(file_path)

# Display the consolidated data
print(consolidated_data.head())
